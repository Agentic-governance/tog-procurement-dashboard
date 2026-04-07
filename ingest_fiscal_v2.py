#!/usr/bin/env python3
"""Parse MIC municipal fiscal data (歳入内訳) and store in DB."""
import sqlite3
import json
import openpyxl
from datetime import datetime

DB = "/app/data/monitor.db"
XLSX = "/tmp/muni_fiscal_r5.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

conn = sqlite3.connect(DB)
cur = conn.cursor()
cur.execute("""CREATE TABLE IF NOT EXISTS fiscal_data (
    muni_code TEXT NOT NULL,
    fiscal_year TEXT NOT NULL,
    muni_name TEXT,
    local_tax INTEGER,
    resident_tax_individual INTEGER,
    resident_tax_corporate INTEGER,
    raw_data TEXT,
    updated_at TEXT,
    PRIMARY KEY (muni_code, fiscal_year)
)""")

inserted = 0
for row_num in range(17, ws.max_row + 1):
    code = ws.cell(row=row_num, column=15).value
    name = ws.cell(row=row_num, column=16).value
    tax = ws.cell(row=row_num, column=17).value
    tax_ind = ws.cell(row=row_num, column=18).value
    tax_corp = ws.cell(row=row_num, column=19).value

    if code and str(code).strip().isdigit() and len(str(code).strip()) == 6:
        code = str(code).strip()
        name = str(name or "").strip()
        try:
            cur.execute("""INSERT OR REPLACE INTO fiscal_data
                (muni_code, fiscal_year, muni_name, local_tax, resident_tax_individual, resident_tax_corporate, raw_data, updated_at)
                VALUES (?, 'R5', ?, ?, ?, ?, ?, ?)""",
                (code, name,
                 int(tax) if tax else None,
                 int(tax_ind) if tax_ind else None,
                 int(tax_corp) if tax_corp else None,
                 json.dumps({"source": "soumu_r5_kessan", "unit": "千円"}, ensure_ascii=False),
                 datetime.now().isoformat()))
            inserted += 1
        except Exception as e:
            pass

conn.commit()
wb.close()

cur.execute("SELECT COUNT(*) FROM fiscal_data")
print(f"Inserted: {inserted}, Total fiscal rows: {cur.fetchone()[0]}", flush=True)

# Sample data
cur.execute("SELECT muni_code, muni_name, local_tax, resident_tax_individual, resident_tax_corporate FROM fiscal_data LIMIT 10")
print("\nSample data (unit: 千円):")
for code, name, tax, ind, corp in cur.fetchall():
    print(f"  {code} {name}: 地方税={tax}, 個人住民税={ind}, 法人住民税={corp}")
conn.close()
